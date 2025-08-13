"""
Korean VC Proposal Automation Platform for Government RFPs
Tailored for 2025 KIF (Korean Information and Communications Fund) GP Selection
"""

import streamlit as st
import pandas as pd
import json
import hashlib
import os
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
import re
from io import BytesIO
import tempfile
import shutil

# Database imports
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime, ForeignKey, Float
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship, Session
from sqlalchemy.sql import func

# PDF and Excel handling
import PyPDF2
from pdfplumber import PDF
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Initialize database
Base = declarative_base()
engine = create_engine('sqlite:///vc_proposal_platform.db', echo=False)
SessionLocal = sessionmaker(bind=engine)

# Database Models
class User(Base):
    __tablename__ = 'users'
    
    id = Column(Integer, primary_key=True)
    username = Column(String(100), unique=True, nullable=False)
    password_hash = Column(String(256), nullable=False)
    firm_name = Column(String(200), nullable=False)
    created_at = Column(DateTime, default=func.now())
    
    proposal_data = relationship("ProposalData", back_populates="user")

class ProposalData(Base):
    __tablename__ = 'proposal_data'
    
    id = Column(Integer, primary_key=True)
    user_id = Column(Integer, ForeignKey('users.id'), nullable=False)
    sheet_id = Column(String(50), nullable=False)
    data_json = Column(Text, nullable=False)
    version = Column(String(100), default='base')
    created_at = Column(DateTime, default=func.now())
    updated_at = Column(DateTime, default=func.now(), onupdate=func.now())
    
    user = relationship("User", back_populates="proposal_data")

# Create tables
Base.metadata.create_all(engine)

# Sheet configurations for 2025 KIF (Real Template Structure)
SHEET_CONFIG = {
    "표지": {"reusability": "low", "category": "기본정보", "description": "Cover Page"},
    "1-0.제안펀드 구성": {"reusability": "low", "category": "펀드구성", "description": "Fund Composition"},
    "1-1.펀드체계 제안": {"reusability": "low", "category": "펀드구성", "description": "Fund Framework Proposal"},
    "1-2.재무실적": {"reusability": "high", "category": "재무정보", "description": "Financial Performance"},
    "1-3.준법성": {"reusability": "high", "category": "컴플라이언스", "description": "Compliance"},
    "1-4.핵심운용인력 관리현황": {"reusability": "high", "category": "인력정보", "description": "Core Operating Personnel"},
    "1-5.조합 결성내역": {"reusability": "medium", "category": "펀드구성", "description": "Fund Formation Details"},
    "2-1.청산펀드 총괄": {"reusability": "high", "category": "운용실적", "description": "Liquidation Fund Overview"},
    "2-1-1.청산펀드 세부1": {"reusability": "high", "category": "운용실적", "description": "Liquidation Fund Details 1"},
    "2-1-2.청산펀드 세부2": {"reusability": "high", "category": "운용실적", "description": "Liquidation Fund Details 2"},
    "2-2.운용중인 펀드 총괄": {"reusability": "high", "category": "운용실적", "description": "Active Fund Overview"},
    "2-2-1.운용펀드 세부1": {"reusability": "high", "category": "운용실적", "description": "Active Fund Details 1"},
    "2-2-2.운용펀드 세부2": {"reusability": "high", "category": "운용실적", "description": "Active Fund Details 2"},
    "2-3.KIF 펀드 운용실적": {"reusability": "high", "category": "운용실적", "description": "KIF Fund Performance"},
    "2-4.본계정 투자내역": {"reusability": "high", "category": "투자실적", "description": "Main Account Investments"},
    "3-1.핵심운용인력 경력기간": {"reusability": "high", "category": "인력정보", "description": "Personnel Career Periods"},
    "3-2.개별 투자실적1": {"reusability": "high", "category": "인력정보", "description": "Individual Investment Performance 1"},
    "3-3.개별 투자실적2": {"reusability": "high", "category": "인력정보", "description": "Individual Investment Performance 2"},
    "3-4.개별 투자실적3": {"reusability": "high", "category": "인력정보", "description": "Individual Investment Performance 3"}
}

# Utility Functions
def hash_password(password: str) -> str:
    """Hash password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash"""
    return hash_password(password) == password_hash

def parse_rfp_pdf(pdf_path: str) -> Dict[str, Any]:
    """Extract key RFP information from PDF with 2025 KIF specific patterns"""
    rfp_info = {
        'announcement_date': '',
        'submission_deadline': '',
        'total_fund_size': '',
        'fund_count': '',
        'investment_areas': [],
        'mandatory_investment': '',
        'fund_duration': '',
        'gp_contribution': '',
        'core_personnel_requirements': {},
        'evaluation_process': [],
        'exclusion_criteria': [],
        'kif_specific_requirements': []
    }
    
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            full_text = ""
            
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                full_text += page.extract_text()
            
            # Extract announcement date (공고일)
            announce_pattern = r'공고일\s*[:：]?\s*(\d{4})[.\-년]\s*(\d{1,2})[.\-월]\s*(\d{1,2})'
            announce_match = re.search(announce_pattern, full_text)
            if announce_match:
                rfp_info['announcement_date'] = f"{announce_match.group(1)}-{announce_match.group(2):0>2}-{announce_match.group(3):0>2}"
            
            # Extract submission deadline (접수마감)
            deadline_pattern = r'접수마감\s*[:：]?\s*•?\s*(\d{4})[.\-년]?\s*(\d{1,2})\s*월?\s*(\d{1,2})\s*일?'
            deadline_match = re.search(deadline_pattern, full_text)
            if deadline_match:
                rfp_info['submission_deadline'] = f"{deadline_match.group(1)}-{deadline_match.group(2):0>2}-{deadline_match.group(3):0>2}"
            
            # Extract total fund size (출자규모)
            total_size_pattern = r'출자규모\s*[:：]?\s*•?\s*([\d,]+)\s*억'
            total_size_match = re.search(total_size_pattern, full_text)
            if total_size_match:
                rfp_info['total_fund_size'] = total_size_match.group(1).replace(',', '') + '억원'
            
            # Extract fund count (조합 수)
            fund_count_pattern = r'조\s*합\s*수\s*[:：]?\s*•?\s*(\d+)\s*개'
            fund_count_match = re.search(fund_count_pattern, full_text)
            if fund_count_match:
                rfp_info['fund_count'] = fund_count_match.group(1) + '개'
            
            # Extract investment areas
            investment_areas = []
            if 'AI·AX 혁신' in full_text:
                investment_areas.append('AI·AX 혁신')
            if 'AI·ICT' in full_text:
                investment_areas.append('AI·ICT')
            if 'ICT 기술사업화' in full_text:
                investment_areas.append('ICT 기술사업화')
            if 'AI 반도체' in full_text:
                investment_areas.append('AI 반도체')
            rfp_info['investment_areas'] = investment_areas
            
            # Extract mandatory investment ratio (의무투자)
            mandatory_pattern = r'의무투자\s*금액\s*[:：]?\s*.*?(\d+)%\s*이상'
            mandatory_match = re.search(mandatory_pattern, full_text)
            if mandatory_match:
                rfp_info['mandatory_investment'] = mandatory_match.group(1) + '%'
            
            # Extract fund duration (존속기간)
            duration_pattern = r'존속\s*기간\s*[:：]?\s*∙?\s*(\d+)\s*년\s*이내'
            duration_match = re.search(duration_pattern, full_text)
            if duration_match:
                rfp_info['fund_duration'] = duration_match.group(1) + '년 이내'
            
            # Extract GP contribution ratio (운용사 출자비율)
            gp_contrib_pattern = r'운용사\s*출자비율\s*[:：]?\s*∙?\s*약정총액의\s*(\d+)%\s*이상'
            gp_contrib_match = re.search(gp_contrib_pattern, full_text)
            if gp_contrib_match:
                rfp_info['gp_contribution'] = '약정총액의 ' + gp_contrib_match.group(1) + '% 이상'
            
            # Extract core personnel requirements (핵심운용인력)
            if '핵심운용인력' in full_text:
                personnel_reqs = {}
                if '총3인이상' in full_text:
                    personnel_reqs['minimum_count'] = '3인 이상'
                elif '2인이상' in full_text:
                    personnel_reqs['minimum_count'] = '2인 이상 (200억원 이하 펀드)'
                
                if '대표펀드매니저는 5년 이상' in full_text:
                    personnel_reqs['lead_manager_experience'] = '5년 이상'
                if '기타 핵심운용인력은3년이상' in full_text:
                    personnel_reqs['other_experience'] = '3년 이상'
                
                rfp_info['core_personnel_requirements'] = personnel_reqs
            
            # Extract evaluation process (선정절차)
            eval_process = []
            if '1차심의(서류평가)' in full_text:
                eval_process.append('1차 심의 (서류평가)')
            if '현장실사' in full_text:
                eval_process.append('현장실사')
            if '2차심의(PT발표평가)' in full_text:
                eval_process.append('2차 심의 (PT발표평가)')
            if '최종선정' in full_text:
                eval_process.append('최종선정 (우선협상대상자)')
            rfp_info['evaluation_process'] = eval_process
            
            # Extract exclusion criteria (선정배제대상)
            exclusions = []
            if '투자비율이60%미만' in full_text:
                exclusions.append('기존 KIF 펀드 투자비율 60% 미만')
            if '2년이미경과' in full_text:
                exclusions.append('최근 선정 후 2년 미경과')
            if '자본잠식률50%이상' in full_text:
                exclusions.append('자본잠식률 50% 이상')
            if '감봉 이상의 제재' in full_text:
                exclusions.append('대표펀드매니저 제재 이력 (3년 이내)')
            rfp_info['exclusion_criteria'] = exclusions
            
            # Extract KIF specific requirements
            kif_requirements = []
            if 'KIF ERP시스템 의무 사용' in full_text:
                kif_requirements.append('KIF ERP 시스템 의무 사용')
            if '수탁기관' in full_text:
                kif_requirements.append('KIF 지정 수탁기관 사용')
            if '회계감사인' in full_text:
                kif_requirements.append('KIF 지정 조건 만족 회계감사인')
            if '분야별 중복지원 불가' in full_text:
                kif_requirements.append('분야별 중복지원 불가')
            rfp_info['kif_specific_requirements'] = kif_requirements
                
    except Exception as e:
        st.error(f"PDF 파싱 오류: {str(e)}")
    
    return rfp_info

def parse_excel_template(excel_path: str) -> Dict[str, Dict]:
    """Load Excel template and extract structure with comprehensive field detection"""
    template_structure = {}
    
    try:
        wb = load_workbook(excel_path, data_only=False)
        
        # First pass: identify all sheets and their basic structure
        all_sheets_info = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_sheets_info[sheet_name] = {
                'max_row': ws.max_row,
                'max_col': ws.max_column,
                'has_data': any(ws.cell(row=r, column=c).value for r in range(1, min(21, ws.max_row+1)) 
                               for c in range(1, min(11, ws.max_column+1)))
            }
        
        # Enhanced field detection patterns
        field_patterns = {
            'financial': ['자산', '자본', '매출', '이익', '부채', '자본금', '잉여금', '현금', '투자', '손익', '수익', '비용', '감가상각'],
            'personnel': ['성명', '직위', '경력', '학력', '자격', '담당', '인원', '조직', '부서', '팀'],
            'fund': ['펀드', '규모', '기간', '수익률', '배수', 'IRR', 'TVPI', 'DPI', '투자금액', '회수금액'],
            'strategy': ['전략', '계획', '목표', '분야', '섹터', '단계', '정책', '방향', '포트폴리오'],
            'compliance': ['컴플라이언스', '리스크', '관리', '체계', '절차', '제재', '소송', '분쟁'],
            'fees': ['보수', '수수료', '비용', '요율', '관리보수', '성과보수', '운용보수'],
            'dates': ['일자', '날짜', '기간', '년도', '월', '일', '시점', '기준일'],
            'amounts': ['금액', '원', '억원', '백만원', '천원', '달러', '규모', '가치', '평가액']
        }
        
        # Process each sheet with enhanced detection
        for sheet_name in wb.sheetnames:
            if all_sheets_info[sheet_name]['has_data']:
                ws = wb[sheet_name]
                sheet_info = {
                    'max_row': ws.max_row,
                    'max_col': ws.max_column,
                    'fields': {},
                    'formulas': {},
                    'data_cells': {},
                    'field_types': {},
                    'merged_cells': []
                }
                
                # Detect merged cells
                for merged_range in ws.merged_cells.ranges:
                    sheet_info['merged_cells'].append(str(merged_range))
                
                # Comprehensive field scanning
                for row in range(1, min(ws.max_row + 1, 150)):
                    for col in range(1, min(ws.max_column + 1, 30)):
                        cell = ws.cell(row=row, column=col)
                        cell_addr = f"{get_column_letter(col)}{row}"
                        
                        if cell.value is not None:
                            cell_value = str(cell.value).strip()
                            
                            # Formulas
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                sheet_info['formulas'][cell_addr] = cell.value
                            
                            # Field labels (Korean text patterns)
                            elif isinstance(cell.value, str) and len(cell_value) > 0:
                                # Detect field types
                                detected_type = 'general'
                                for field_type, patterns in field_patterns.items():
                                    if any(pattern in cell_value for pattern in patterns):
                                        detected_type = field_type
                                        break
                                
                                # Store field information
                                if len(cell_value) < 100 and any(ord(c) > 127 for c in cell_value):  # Contains Korean
                                    sheet_info['fields'][cell_addr] = {
                                        'label': cell_value,
                                        'type': detected_type,
                                        'row': row,
                                        'col': col
                                    }
                                    sheet_info['field_types'][cell_addr] = detected_type
                            
                            # Data cells (numbers, dates)
                            elif isinstance(cell.value, (int, float)):
                                sheet_info['data_cells'][cell_addr] = {
                                    'value': cell.value,
                                    'type': 'numeric',
                                    'row': row,
                                    'col': col
                                }
                            elif isinstance(cell.value, datetime):
                                sheet_info['data_cells'][cell_addr] = {
                                    'value': cell.value,
                                    'type': 'date',
                                    'row': row,
                                    'col': col
                                }
                
                # Try to match with known sheet configurations
                matched_config = None
                for config_name in SHEET_CONFIG.keys():
                    if config_name in sheet_name or any(key in sheet_name for key in config_name.split('.')):
                        matched_config = config_name
                        break
                
                sheet_info['matched_config'] = matched_config
                sheet_info['sheet_category'] = SHEET_CONFIG.get(matched_config, {}).get('category', 'unknown')
                
                template_structure[sheet_name] = sheet_info
        
        wb.close()
        
        # Store detected structure for debugging
        st.session_state.template_analysis = {
            'total_sheets': len(wb.sheetnames),
            'data_sheets': len(template_structure),
            'field_count': sum(len(sheet['fields']) for sheet in template_structure.values()),
            'formula_count': sum(len(sheet['formulas']) for sheet in template_structure.values())
        }
        
    except Exception as e:
        st.error(f"Excel 템플릿 파싱 오류: {str(e)}")
        st.exception(e)
    
    return template_structure

def load_stored_data(user_id: int) -> Dict[str, Any]:
    """Fetch stored data from database"""
    stored_data = {}
    
    with SessionLocal() as session:
        data_records = session.query(ProposalData).filter_by(user_id=user_id).all()
        
        for record in data_records:
            if record.sheet_id not in stored_data:
                stored_data[record.sheet_id] = {}
            
            try:
                data = json.loads(record.data_json)
                stored_data[record.sheet_id][record.version] = data
            except json.JSONDecodeError:
                continue
    
    return stored_data

def compare_data(stored: Dict, rfp_reqs: Dict, template: Dict) -> Dict[str, Any]:
    """Analyze data availability and suggest improvements"""
    comparison = {
        'available': [],
        'partial': [],
        'missing': [],
        'suggestions': []
    }
    
    # Check each sheet
    for sheet_name in SHEET_CONFIG.keys():
        if sheet_name in stored and stored[sheet_name]:
            # Check if data is complete
            if template.get(sheet_name, {}).get('fields'):
                required_fields = len(template[sheet_name]['fields'])
                if stored[sheet_name].get('base'):
                    filled_fields = len([v for v in stored[sheet_name]['base'].values() if v])
                    if filled_fields >= required_fields * 0.8:
                        comparison['available'].append(sheet_name)
                    elif filled_fields >= required_fields * 0.3:
                        comparison['partial'].append(sheet_name)
                    else:
                        comparison['missing'].append(sheet_name)
                else:
                    comparison['missing'].append(sheet_name)
            else:
                comparison['available'].append(sheet_name)
        else:
            comparison['missing'].append(sheet_name)
    
    # Generate suggestions based on RFP requirements
    if rfp_reqs.get('mandatory_investment'):
        if 'AI' in rfp_reqs['mandatory_investment'] or '인공지능' in rfp_reqs['mandatory_investment']:
            comparison['suggestions'].append("AI/인공지능 투자 전략을 '2-4.투자전략 및 계획'에 추가 필요")
    
    if rfp_reqs.get('sectors'):
        for sector in rfp_reqs['sectors'][:3]:
            comparison['suggestions'].append(f"{sector} 분야 투자 실적을 '2-3.주요펀드 운용 실적'에 강조")
    
    if len(comparison['missing']) > 5:
        comparison['suggestions'].append("주요 데이터가 많이 누락됨. 기본 정보부터 순차적으로 입력 권장")
    
    return comparison

def update_data(user_id: int, sheet_id: str, new_data: Dict, version: str = 'base'):
    """Save or update data in database"""
    with SessionLocal() as session:
        # Check if record exists
        existing = session.query(ProposalData).filter_by(
            user_id=user_id, 
            sheet_id=sheet_id,
            version=version
        ).first()
        
        data_json = json.dumps(new_data, ensure_ascii=False)
        
        if existing:
            existing.data_json = data_json
            existing.updated_at = datetime.now()
        else:
            new_record = ProposalData(
                user_id=user_id,
                sheet_id=sheet_id,
                data_json=data_json,
                version=version
            )
            session.add(new_record)
        
        session.commit()

def generate_filled_excel(template_path: str, stored_data: Dict) -> str:
    """Fill Excel template with stored data"""
    output_path = None
    
    try:
        # Create temp file for output
        temp_dir = tempfile.mkdtemp()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(temp_dir, f"filled_proposal_{timestamp}.xlsx")
        
        # Copy template to output
        shutil.copy2(template_path, output_path)
        
        # Load workbook
        wb = load_workbook(output_path)
        
        # Fill each sheet
        for sheet_name, sheet_data in stored_data.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get the latest version data
                data_to_fill = sheet_data.get('base', {})
                if '2025 KIF Version' in sheet_data:
                    data_to_fill.update(sheet_data['2025 KIF Version'])
                
                # Fill cells based on stored data
                for cell_ref, value in data_to_fill.items():
                    if re.match(r'^[A-Z]+\d+$', cell_ref):  # Valid cell reference
                        try:
                            cell = ws[cell_ref]
                            # Preserve formulas
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                # Handle different data types
                                if isinstance(value, (int, float)):
                                    cell.value = value
                                elif isinstance(value, str):
                                    # Handle date format
                                    if re.match(r'\d{4}\.\d{2}\.\d{2}', value):
                                        cell.value = value
                                    else:
                                        cell.value = value
                                elif value is not None:
                                    cell.value = str(value)
                        except Exception as e:
                            continue
                
                # Special handling for specific sheets
                if sheet_name == "1-2.재무실적":
                    # Fill financial data with proper formatting
                    financial_fields = {
                        'B8': '유동자산', 'C8': '비유동자산', 'D8': '자산총계',
                        'B9': '유동부채', 'C9': '비유동부채', 'D9': '부채총계',
                        'B10': '자본금', 'C10': '자본잉여금', 'D10': '자본총계',
                        'B11': '매출액', 'C11': '영업이익', 'D11': '당기순이익'
                    }
                    
                    for cell_ref, field_name in financial_fields.items():
                        if field_name in data_to_fill:
                            ws[cell_ref] = data_to_fill[field_name]
        
        # Save workbook
        wb.save(output_path)
        wb.close()
        
        return output_path
        
    except Exception as e:
        st.error(f"Excel 생성 오류: {str(e)}")
        if output_path and os.path.exists(output_path):
            os.remove(output_path)
        return None

def validate_input(data: Dict, sheet_id: str) -> List[str]:
    """Validate input data based on 2025 KIF requirements"""
    errors = []
    
    # KIF-specific validation rules
    for field, value in data.items():
        # No empty cells rule (KIF requirement)
        if value is None or str(value).strip() == "":
            errors.append(f"{field}: 빈 셀 불허 (0 또는 해당 데이터 입력 필수)")
        
        # Date format validation (YYYY-MM-DD for KIF)
        if any(keyword in field for keyword in ['일자', '날짜', '기간', '년도']):
            if value and not re.match(r'^\d{4}-\d{2}-\d{2}$', str(value)):
                errors.append(f"{field}: KIF 날짜 형식은 YYYY-MM-DD")
        
        # Percentage validation (첫째 자리까지)
        if '비율' in field or '%' in field or 'IRR' in field:
            if value:
                try:
                    float_val = float(str(value).replace('%', ''))
                    if abs(float_val - round(float_val, 1)) > 0.01:
                        errors.append(f"{field}: 퍼센트는 소수점 첫째 자리까지만 입력")
                except ValueError:
                    errors.append(f"{field}: 유효한 퍼센트 값이 아님")
        
        # Amount validation (정확한 금액 필요)
        if any(keyword in field for keyword in ['금액', '규모', '자산', '자본', '매출', '투자']):
            if value and value != 0:
                try:
                    float(str(value).replace(',', ''))
                except ValueError:
                    errors.append(f"{field}: 유효한 금액이 아님")
        
        # Company/Fund name validation (정식 명칭)
        if any(keyword in field for keyword in ['회사명', '펀드명', '법인명']):
            if value and len(str(value)) < 2:
                errors.append(f"{field}: 정식 명칭 입력 필요")
    
    # Sheet-specific KIF validations
    if sheet_id == "1-2.재무실적":
        # Financial data completeness check
        required_financial = ['자산', '부채', '자본', '매출']
        for req_field in required_financial:
            if not any(req_field in key for key in data.keys()):
                errors.append(f"재무실적: {req_field} 데이터 필수")
    
    elif sheet_id == "1-4.핵심운용인력 관리현황":
        # Core personnel requirements
        if not any('대표' in str(value) or 'CEO' in str(value) for value in data.values()):
            errors.append("핵심운용인력: 대표이사/CEO 정보 필수")
        
        # Career length validation
        for key, value in data.items():
            if '경력' in key and value:
                try:
                    career_years = float(value)
                    if career_years < 0 or career_years > 50:
                        errors.append(f"{key}: 경력년수는 0-50년 범위")
                except ValueError:
                    pass
    
    elif sheet_id == "2-3.KIF 펀드 운용실적":
        # KIF fund performance requirements
        required_metrics = ['수익률', 'IRR', 'TVPI']
        for metric in required_metrics:
            if not any(metric in key for key in data.keys()):
                errors.append(f"KIF 펀드 실적: {metric} 데이터 필수")
    
    elif sheet_id == "1-3.준법성":
        # Compliance requirements
        if not any('리스크' in key or '컴플라이언스' in key for key in data.keys()):
            errors.append("준법성: 리스크 관리 체계 정보 필수")
    
    # A4 print requirement warning
    if len(str(data)) > 10000:  # Rough estimate
        errors.append("주의: 데이터가 A4 인쇄 크기를 초과할 수 있음")
    
    return errors

# Streamlit UI Components
def init_session_state():
    """Initialize session state variables"""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'user_id' not in st.session_state:
        st.session_state.user_id = None
    if 'username' not in st.session_state:
        st.session_state.username = None
    if 'firm_name' not in st.session_state:
        st.session_state.firm_name = None
    if 'uploaded_rfp' not in st.session_state:
        st.session_state.uploaded_rfp = None
    if 'uploaded_template' not in st.session_state:
        st.session_state.uploaded_template = None
    if 'rfp_info' not in st.session_state:
        st.session_state.rfp_info = {}
    if 'template_structure' not in st.session_state:
        st.session_state.template_structure = {}

def login_page():
    """Display login page"""
    st.title("🏢 한국 VC 제안서 자동화 플랫폼")
    st.subheader("2025 KIF GP 선정 최적화 시스템")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("### 로그인")
        
        with st.form("login_form"):
            username = st.text_input("사용자명")
            password = st.text_input("비밀번호", type="password")
            firm_name = st.text_input("운용사명 (신규 등록시)")
            
            col_login, col_register = st.columns(2)
            
            with col_login:
                login_btn = st.form_submit_button("로그인", use_container_width=True)
            with col_register:
                register_btn = st.form_submit_button("신규 등록", use_container_width=True)
            
            if login_btn:
                with SessionLocal() as session:
                    user = session.query(User).filter_by(username=username).first()
                    if user and verify_password(password, user.password_hash):
                        st.session_state.authenticated = True
                        st.session_state.user_id = user.id
                        st.session_state.username = user.username
                        st.session_state.firm_name = user.firm_name
                        st.success(f"환영합니다, {user.firm_name}!")
                        st.rerun()
                    else:
                        st.error("잘못된 사용자명 또는 비밀번호")
            
            if register_btn:
                if username and password and firm_name:
                    with SessionLocal() as session:
                        existing = session.query(User).filter_by(username=username).first()
                        if existing:
                            st.error("이미 존재하는 사용자명")
                        else:
                            new_user = User(
                                username=username,
                                password_hash=hash_password(password),
                                firm_name=firm_name
                            )
                            session.add(new_user)
                            session.commit()
                            st.success("등록 완료! 로그인해주세요.")
                else:
                    st.error("모든 필드를 입력해주세요")

def sidebar_menu():
    """Display sidebar menu"""
    with st.sidebar:
        st.title(f"🏢 {st.session_state.firm_name}")
        st.divider()
        
        st.markdown("### 📁 파일 업로드")
        
        # RFP PDF Upload
        rfp_file = st.file_uploader(
            "RFP 공고문 (PDF)",
            type=['pdf'],
            key="rfp_uploader"
        )
        
        if rfp_file:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(rfp_file.read())
                st.session_state.uploaded_rfp = tmp_file.name
                st.session_state.rfp_info = parse_rfp_pdf(tmp_file.name)
                st.success("RFP 파싱 완료")
        
        # Excel Template Upload
        template_file = st.file_uploader(
            "제출 양식 (Excel)",
            type=['xlsx', 'xls'],
            key="template_uploader"
        )
        
        if template_file:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(template_file.read())
                st.session_state.uploaded_template = tmp_file.name
                st.session_state.template_structure = parse_excel_template(tmp_file.name)
                st.success("템플릿 파싱 완료")
        
        st.divider()
        
        if st.button("🚪 로그아웃", use_container_width=True):
            for key in st.session_state.keys():
                del st.session_state[key]
            st.rerun()

def data_vault_tab():
    """Display data vault dashboard"""
    st.header("📊 데이터 보관소")
    
    # Load stored data
    stored_data = load_stored_data(st.session_state.user_id)
    
    # Create status overview
    col1, col2, col3 = st.columns(3)
    
    available_sheets = []
    partial_sheets = []
    missing_sheets = []
    
    for sheet_name in SHEET_CONFIG.keys():
        if sheet_name in stored_data and stored_data[sheet_name]:
            if any(stored_data[sheet_name].values()):
                available_sheets.append(sheet_name)
            else:
                partial_sheets.append(sheet_name)
        else:
            missing_sheets.append(sheet_name)
    
    with col1:
        st.metric("✅ 완료", len(available_sheets))
    with col2:
        st.metric("⚠️ 부분 완료", len(partial_sheets))
    with col3:
        st.metric("❌ 미입력", len(missing_sheets))
    
    st.divider()
    
    # Display sheet status with categories
    categories = {}
    for sheet_name, config in SHEET_CONFIG.items():
        category = config['category']
        if category not in categories:
            categories[category] = []
        categories[category].append(sheet_name)
    
    for category, sheets in categories.items():
        with st.expander(f"📁 {category}", expanded=True):
            for sheet in sheets:
                col1, col2, col3 = st.columns([3, 1, 1])
                
                with col1:
                    if sheet in available_sheets:
                        st.markdown(f"✅ **{sheet}**")
                    elif sheet in partial_sheets:
                        st.markdown(f"⚠️ **{sheet}**")
                    else:
                        st.markdown(f"❌ **{sheet}**")
                
                with col2:
                    st.markdown(f"*{SHEET_CONFIG[sheet]['reusability']} 재사용성*")
                
                with col3:
                    if st.button("편집", key=f"edit_{sheet}"):
                        st.session_state.editing_sheet = sheet

def input_forms_tab():
    """Display input forms for data entry"""
    st.header("✏️ 데이터 입력/수정")
    
    # Sheet selector
    sheet_to_edit = st.selectbox(
        "편집할 시트 선택",
        options=list(SHEET_CONFIG.keys()),
        key="sheet_selector"
    )
    
    if sheet_to_edit:
        st.subheader(f"📝 {sheet_to_edit}")
        
        # Load existing data
        stored_data = load_stored_data(st.session_state.user_id)
        existing_data = stored_data.get(sheet_to_edit, {}).get('base', {})
        
        # Create dynamic form based on sheet type
        with st.form(f"form_{sheet_to_edit}"):
            form_data = {}
            
            # Sheet-specific forms
            if sheet_to_edit == "1-2.재무실적":
                st.markdown("### 재무 정보 (단위: 백만원)")
                
                years = st.columns(4)
                year_labels = ["2021년", "2022년", "2023년", "2024년(예상)"]
                
                for idx, year_col in enumerate(years):
                    with year_col:
                        st.markdown(f"**{year_labels[idx]}**")
                        form_data[f"유동자산_{year_labels[idx]}"] = st.number_input(
                            "유동자산", 
                            value=float(existing_data.get(f"유동자산_{year_labels[idx]}", 0)),
                            key=f"유동자산_{idx}"
                        )
                        form_data[f"비유동자산_{year_labels[idx]}"] = st.number_input(
                            "비유동자산",
                            value=float(existing_data.get(f"비유동자산_{year_labels[idx]}", 0)),
                            key=f"비유동자산_{idx}"
                        )
                        form_data[f"매출액_{year_labels[idx]}"] = st.number_input(
                            "매출액",
                            value=float(existing_data.get(f"매출액_{year_labels[idx]}", 0)),
                            key=f"매출액_{idx}"
                        )
                        form_data[f"영업이익_{year_labels[idx]}"] = st.number_input(
                            "영업이익",
                            value=float(existing_data.get(f"영업이익_{year_labels[idx]}", 0)),
                            key=f"영업이익_{idx}"
                        )
            
            elif sheet_to_edit == "1-4.핵심운용인력 관리현황":
                st.markdown("### 핵심 운용인력")
                
                num_members = st.number_input("인력 수", min_value=1, max_value=10, value=3)
                
                for i in range(int(num_members)):
                    st.markdown(f"#### 인력 {i+1}")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        form_data[f"성명_{i}"] = st.text_input(
                            "성명",
                            value=existing_data.get(f"성명_{i}", ""),
                            key=f"name_{i}"
                        )
                    with col2:
                        form_data[f"직위_{i}"] = st.text_input(
                            "직위",
                            value=existing_data.get(f"직위_{i}", ""),
                            key=f"position_{i}"
                        )
                    with col3:
                        form_data[f"경력년수_{i}"] = st.number_input(
                            "경력(년)",
                            value=int(existing_data.get(f"경력년수_{i}", 0)),
                            key=f"experience_{i}"
                        )
            
            elif sheet_to_edit == "2-4.투자전략 및 계획":
                st.markdown("### 투자 전략")
                
                form_data["주요투자분야"] = st.multiselect(
                    "주요 투자 분야",
                    options=["AI/인공지능", "5G/6G", "블록체인", "메타버스", "바이오", 
                            "헬스케어", "반도체", "배터리", "모빌리티", "로봇", "우주"],
                    default=existing_data.get("주요투자분야", [])
                )
                
                form_data["투자전략"] = st.text_area(
                    "투자 전략 설명",
                    value=existing_data.get("투자전략", ""),
                    height=200
                )
                
                form_data["목표수익률"] = st.number_input(
                    "목표 수익률 (%)",
                    value=float(existing_data.get("목표수익률", 0)),
                    min_value=0.0,
                    max_value=100.0
                )
            
            else:
                # Generic form for other sheets
                st.info(f"{sheet_to_edit}에 대한 일반 입력 폼")
                
                # Add some generic fields
                form_data["field1"] = st.text_input(
                    "필드 1",
                    value=existing_data.get("field1", "")
                )
                form_data["field2"] = st.text_input(
                    "필드 2", 
                    value=existing_data.get("field2", "")
                )
                form_data["field3"] = st.number_input(
                    "숫자 필드",
                    value=float(existing_data.get("field3", 0))
                )
            
            # Version selection
            version = st.selectbox(
                "버전",
                options=["base", "2025 KIF Version", "Custom"],
                index=0
            )
            
            # Submit button
            submitted = st.form_submit_button("💾 저장", use_container_width=True)
            
            if submitted:
                # Validate input
                errors = validate_input(form_data, sheet_to_edit)
                
                if errors:
                    for error in errors:
                        st.error(error)
                else:
                    # Save to database
                    update_data(st.session_state.user_id, sheet_to_edit, form_data, version)
                    st.success(f"{sheet_to_edit} 데이터 저장 완료!")
                    st.balloons()

def analysis_tab():
    """Display analysis and suggestions"""
    st.header("🔍 분석 및 제안")
    
    if not st.session_state.rfp_info and not st.session_state.template_structure:
        st.warning("먼저 RFP PDF와 Excel 템플릿을 업로드해주세요.")
        return
    
    # Load and analyze data
    stored_data = load_stored_data(st.session_state.user_id)
    comparison = compare_data(
        stored_data, 
        st.session_state.rfp_info,
        st.session_state.template_structure
    )
    
    # Display RFP requirements
    if st.session_state.rfp_info:
        st.subheader("📋 2025 KIF GP 선정 주요 요구사항")
        
        # Basic info
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.session_state.rfp_info.get('submission_deadline'):
                st.error(f"**⏰ 접수마감**: {st.session_state.rfp_info['submission_deadline']}")
            if st.session_state.rfp_info.get('total_fund_size'):
                st.info(f"**💰 총 출자규모**: {st.session_state.rfp_info['total_fund_size']}")
        
        with col2:
            if st.session_state.rfp_info.get('fund_count'):
                st.info(f"**📊 조합 수**: {st.session_state.rfp_info['fund_count']}")
            if st.session_state.rfp_info.get('fund_duration'):
                st.info(f"**⏳ 존속기간**: {st.session_state.rfp_info['fund_duration']}")
        
        with col3:
            if st.session_state.rfp_info.get('mandatory_investment'):
                st.warning(f"**🎯 의무투자**: {st.session_state.rfp_info['mandatory_investment']} 이상")
            if st.session_state.rfp_info.get('gp_contribution'):
                st.info(f"**🏢 GP 출자**: {st.session_state.rfp_info['gp_contribution']}")
        
        # Investment areas
        if st.session_state.rfp_info.get('investment_areas'):
            st.subheader("🎯 투자 분야")
            areas_text = " | ".join(st.session_state.rfp_info['investment_areas'])
            st.success(f"**선택 가능 분야**: {areas_text}")
        
        # Core personnel requirements
        if st.session_state.rfp_info.get('core_personnel_requirements'):
            st.subheader("👥 핵심운용인력 요구사항")
            personnel_reqs = st.session_state.rfp_info['core_personnel_requirements']
            
            col1, col2 = st.columns(2)
            with col1:
                if personnel_reqs.get('minimum_count'):
                    st.info(f"**최소 인원**: {personnel_reqs['minimum_count']}")
                if personnel_reqs.get('lead_manager_experience'):
                    st.warning(f"**대표펀드매니저**: {personnel_reqs['lead_manager_experience']} 경력")
            with col2:
                if personnel_reqs.get('other_experience'):
                    st.info(f"**기타 핵심인력**: {personnel_reqs['other_experience']} 경력")
        
        # Evaluation process
        if st.session_state.rfp_info.get('evaluation_process'):
            st.subheader("📝 선정 절차")
            process_steps = " → ".join(st.session_state.rfp_info['evaluation_process'])
            st.info(f"**평가 단계**: {process_steps}")
        
        # Exclusion criteria
        if st.session_state.rfp_info.get('exclusion_criteria'):
            st.subheader("❌ 선정 배제 대상")
            for exclusion in st.session_state.rfp_info['exclusion_criteria']:
                st.error(f"• {exclusion}")
        
        # KIF specific requirements
        if st.session_state.rfp_info.get('kif_specific_requirements'):
            st.subheader("⚙️ KIF 특별 요구사항")
            for requirement in st.session_state.rfp_info['kif_specific_requirements']:
                st.warning(f"• {requirement}")
    
    st.divider()
    
    # Display data completeness
    st.subheader("📊 데이터 완성도 분석")
    
    # Progress bars for each category
    categories_progress = {}
    for sheet_name in SHEET_CONFIG.keys():
        category = SHEET_CONFIG[sheet_name]['category']
        if category not in categories_progress:
            categories_progress[category] = {'complete': 0, 'total': 0}
        
        categories_progress[category]['total'] += 1
        if sheet_name in comparison['available']:
            categories_progress[category]['complete'] += 1
        elif sheet_name in comparison['partial']:
            categories_progress[category]['complete'] += 0.5
    
    for category, progress in categories_progress.items():
        completion_rate = (progress['complete'] / progress['total']) * 100 if progress['total'] > 0 else 0
        st.progress(completion_rate / 100)
        st.caption(f"{category}: {completion_rate:.0f}% 완료")
    
    st.divider()
    
    # Display suggestions
    if comparison['suggestions']:
        st.subheader("💡 개선 제안")
        for suggestion in comparison['suggestions']:
            st.warning(f"• {suggestion}")
    
    # Missing data alert
    if comparison['missing']:
        st.subheader("⚠️ 미입력 항목")
        missing_by_category = {}
        for sheet in comparison['missing']:
            category = SHEET_CONFIG[sheet]['category']
            if category not in missing_by_category:
                missing_by_category[category] = []
            missing_by_category[category].append(sheet)
        
        for category, sheets in missing_by_category.items():
            st.error(f"**{category}**: {', '.join(sheets)}")

def generation_tab():
    """Display Excel generation interface"""
    st.header("📄 제안서 생성")
    
    if not st.session_state.uploaded_template:
        st.warning("먼저 Excel 템플릿을 업로드해주세요.")
        return
    
    # Load stored data
    stored_data = load_stored_data(st.session_state.user_id)
    
    # Data summary
    st.subheader("📊 데이터 요약")
    
    total_sheets = len(SHEET_CONFIG)
    filled_sheets = len([s for s in stored_data.keys() if s in SHEET_CONFIG])
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("전체 시트", total_sheets)
    with col2:
        st.metric("입력 완료", filled_sheets)
    with col3:
        completion_rate = (filled_sheets / total_sheets * 100) if total_sheets > 0 else 0
        st.metric("완성도", f"{completion_rate:.0f}%")
    
    st.divider()
    
    # Generation options
    st.subheader("⚙️ 생성 옵션")
    
    col1, col2 = st.columns(2)
    
    with col1:
        version_to_use = st.selectbox(
            "사용할 데이터 버전",
            options=["base", "2025 KIF Version", "최신 버전 자동 선택"],
            index=2
        )
    
    with col2:
        include_empty = st.checkbox("미입력 시트 포함", value=True)
    
    # Preview section
    with st.expander("👁️ 데이터 미리보기", expanded=False):
        for sheet_name in SHEET_CONFIG.keys():
            if sheet_name in stored_data:
                st.markdown(f"**{sheet_name}**")
                sheet_data = stored_data[sheet_name].get('base', {})
                if sheet_data:
                    preview_data = dict(list(sheet_data.items())[:5])  # Show first 5 items
                    for key, value in preview_data.items():
                        st.text(f"  {key}: {value}")
                else:
                    st.text("  (데이터 없음)")
    
    st.divider()
    
    # Generate button
    if st.button("🚀 Excel 제안서 생성", type="primary", use_container_width=True):
        with st.spinner("제안서 생성 중..."):
            # Generate Excel
            output_path = generate_filled_excel(
                st.session_state.uploaded_template,
                stored_data
            )
            
            if output_path and os.path.exists(output_path):
                # Read file for download
                with open(output_path, 'rb') as f:
                    file_data = f.read()
                
                # Create download button
                st.success("✅ 제안서 생성 완료!")
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="📥 다운로드",
                    data=file_data,
                    file_name=f"KIF_제안서_{st.session_state.firm_name}_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Clean up temp file
                os.remove(output_path)
                
                # Show completion message
                st.balloons()
                st.info("생성된 파일을 다운로드하여 최종 검토 후 제출하세요.")
            else:
                st.error("제안서 생성 실패. 데이터를 확인해주세요.")

def history_tab():
    """Display version history"""
    st.header("📜 버전 관리")
    
    with SessionLocal() as session:
        # Get all proposal data for user
        all_data = session.query(ProposalData).filter_by(
            user_id=st.session_state.user_id
        ).order_by(ProposalData.updated_at.desc()).all()
        
        if all_data:
            # Group by version
            versions = {}
            for record in all_data:
                if record.version not in versions:
                    versions[record.version] = []
                versions[record.version].append(record)
            
            # Display versions
            for version_name, records in versions.items():
                with st.expander(f"📌 {version_name}", expanded=(version_name == "base")):
                    st.markdown(f"**시트 수**: {len(records)}")
                    st.markdown(f"**최종 수정**: {max(r.updated_at for r in records).strftime('%Y-%m-%d %H:%M')}")
                    
                    # List sheets in this version
                    for record in records[:5]:  # Show first 5
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            st.text(f"• {record.sheet_id}")
                        with col2:
                            st.caption(record.updated_at.strftime('%m/%d'))
                    
                    if len(records) > 5:
                        st.caption(f"... 외 {len(records)-5}개 시트")
        else:
            st.info("아직 저장된 데이터가 없습니다.")

def template_analysis_tab():
    """Display template analysis and structure"""
    st.header("🔧 템플릿 구조 분석")
    
    if not st.session_state.template_structure:
        st.warning("먼저 Excel 템플릿을 업로드해주세요.")
        return
    
    # Display analysis summary
    if hasattr(st.session_state, 'template_analysis'):
        analysis = st.session_state.template_analysis
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("총 시트 수", analysis.get('total_sheets', 0))
        with col2:
            st.metric("데이터 시트", analysis.get('data_sheets', 0))
        with col3:
            st.metric("감지된 필드", analysis.get('field_count', 0))
        with col4:
            st.metric("수식 셀", analysis.get('formula_count', 0))
    
    st.divider()
    
    # Display detected sheets and their structure
    st.subheader("📋 감지된 시트 구조")
    
    for sheet_name, sheet_info in st.session_state.template_structure.items():
        with st.expander(f"📄 {sheet_name}", expanded=False):
            
            # Basic info
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("행 수", sheet_info.get('max_row', 0))
            with col2:
                st.metric("열 수", sheet_info.get('max_col', 0))
            with col3:
                matched_config = sheet_info.get('matched_config', '매칭 없음')
                st.info(f"매칭: {matched_config}")
            
            # Fields analysis
            if sheet_info.get('fields'):
                st.markdown("### 📝 감지된 필드")
                
                fields_df = []
                for cell_addr, field_info in sheet_info['fields'].items():
                    if isinstance(field_info, dict):
                        fields_df.append({
                            '셀 주소': cell_addr,
                            '필드명': field_info.get('label', ''),
                            '유형': field_info.get('type', 'general'),
                            '행': field_info.get('row', ''),
                            '열': field_info.get('col', '')
                        })
                    else:
                        fields_df.append({
                            '셀 주소': cell_addr,
                            '필드명': str(field_info),
                            '유형': 'general',
                            '행': '',
                            '열': ''
                        })
                
                if fields_df:
                    df = pd.DataFrame(fields_df)
                    st.dataframe(df, use_container_width=True, height=200)
            
            # Formulas
            if sheet_info.get('formulas'):
                st.markdown("### 🧮 수식 셀")
                formula_data = []
                for cell_addr, formula in list(sheet_info['formulas'].items())[:10]:  # Show first 10
                    formula_data.append({
                        '셀 주소': cell_addr,
                        '수식': formula[:50] + '...' if len(formula) > 50 else formula
                    })
                
                if formula_data:
                    df_formulas = pd.DataFrame(formula_data)
                    st.dataframe(df_formulas, use_container_width=True)
                
                if len(sheet_info['formulas']) > 10:
                    st.caption(f"... 외 {len(sheet_info['formulas']) - 10}개 수식")
            
            # Data cells preview
            if sheet_info.get('data_cells'):
                st.markdown("### 📊 데이터 셀 (샘플)")
                data_cells = list(sheet_info['data_cells'].items())[:5]
                for cell_addr, cell_info in data_cells:
                    st.text(f"{cell_addr}: {cell_info.get('value', '')} ({cell_info.get('type', '')})")
                
                if len(sheet_info['data_cells']) > 5:
                    st.caption(f"... 외 {len(sheet_info['data_cells']) - 5}개 데이터 셀")
            
            # Merged cells
            if sheet_info.get('merged_cells'):
                st.markdown("### 🔗 병합된 셀")
                for merged in sheet_info['merged_cells'][:5]:
                    st.text(f"• {merged}")
                if len(sheet_info['merged_cells']) > 5:
                    st.caption(f"... 외 {len(sheet_info['merged_cells']) - 5}개 병합 셀")
    
    st.divider()
    
    # Field type distribution
    st.subheader("📈 필드 유형 분포")
    
    field_type_count = {}
    for sheet_info in st.session_state.template_structure.values():
        for field_info in sheet_info.get('fields', {}).values():
            if isinstance(field_info, dict):
                field_type = field_info.get('type', 'general')
                field_type_count[field_type] = field_type_count.get(field_type, 0) + 1
    
    if field_type_count:
        # Create a chart
        type_names = list(field_type_count.keys())
        type_counts = list(field_type_count.values())
        
        chart_data = pd.DataFrame({
            '필드 유형': type_names,
            '개수': type_counts
        })
        
        st.bar_chart(data=chart_data.set_index('필드 유형'))
        
        # Show detailed breakdown
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 유형별 상세")
            for field_type, count in field_type_count.items():
                st.metric(field_type, count)
    
    # Template mapping suggestions
    st.divider()
    st.subheader("💡 매핑 제안")
    
    unmatched_sheets = [
        sheet_name for sheet_name, sheet_info in st.session_state.template_structure.items()
        if not sheet_info.get('matched_config')
    ]
    
    if unmatched_sheets:
        st.warning("다음 시트들이 표준 구성과 매칭되지 않았습니다:")
        for sheet in unmatched_sheets:
            st.text(f"• {sheet}")
        
        st.info("이 시트들에 대한 매핑을 수동으로 설정하거나, 표준 시트명과 유사하게 이름을 변경하는 것을 권장합니다.")
    else:
        st.success("모든 시트가 표준 구성과 성공적으로 매칭되었습니다!")
    
    # Export template structure
    if st.button("📥 템플릿 구조 JSON 다운로드"):
        template_json = json.dumps(st.session_state.template_structure, ensure_ascii=False, indent=2)
        st.download_button(
            label="다운로드",
            data=template_json,
            file_name="template_structure.json",
            mime="application/json"
        )

def main():
    """Main application"""
    st.set_page_config(
        page_title="VC 제안서 자동화 플랫폼",
        page_icon="🏢",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-left: 20px;
        padding-right: 20px;
        background-color: #f0f2f6;
        border-radius: 8px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1f77b4;
        color: white;
    }
    div[data-testid="metric-container"] {
        background-color: #f0f2f6;
        border: 1px solid #ddd;
        padding: 10px;
        border-radius: 8px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Initialize session state
    init_session_state()
    
    # Show login if not authenticated
    if not st.session_state.authenticated:
        login_page()
    else:
        # Main app with sidebar
        sidebar_menu()
        
        # Main content area with tabs
        tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
            "📊 데이터 보관소",
            "✏️ 입력/수정", 
            "🔍 분석",
            "📄 생성",
            "📜 히스토리",
            "🔧 템플릿 분석"
        ])
        
        with tab1:
            data_vault_tab()
        
        with tab2:
            input_forms_tab()
        
        with tab3:
            analysis_tab()
        
        with tab4:
            generation_tab()
        
        with tab5:
            history_tab()
        
        with tab6:
            template_analysis_tab()

if __name__ == "__main__":
    main()